import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from tkcalendar import DateEntry
import mysql.connector
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import openpyxl
from datetime import date # Cần import để dùng cho DateEntry mặc định và kiểm tra

# ====== KẾT NỐI MYSQL VÀ KHỞI TẠO CSDL ======
def connect_db():
    """Thiết lập kết nối tới cơ sở dữ liệu qlktx"""
    try:
        # Thay đổi host, user, password nếu cần
        return mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="qlktx"
        )
    except mysql.connector.Error as err:
        messagebox.showerror("Lỗi kết nối", f"Không thể kết nối CSDL: {err}\nVui lòng đảm bảo MySQL đang chạy và CSDL 'qlktx' đã được tạo.")
        sys.exit(1)

def initialize_db():
    """Tạo bảng 'sinhvien' nếu chưa tồn tại, có thêm các cột mới."""
    conn = connect_db()
    try:
        cur = conn.cursor()
        create_table_query = """
        CREATE TABLE IF NOT EXISTS sinhvien (
            maso VARCHAR(20) PRIMARY KEY,
            holot VARCHAR(50) NOT NULL,
            ten VARCHAR(20) NOT NULL,
            gioitinh VARCHAR(5),
            ngaysinh DATE,
            lop VARCHAR(20),
            phong_so VARCHAR(10),
            thanhtien DECIMAL(10, 0) DEFAULT 2000000,
            trangthaidongtien VARCHAR(50) DEFAULT 'Chưa đóng',
            ngayvao DATE
        )
        """
        cur.execute(create_table_query)
        conn.commit()
    except Exception as e:
        messagebox.showerror("Lỗi khởi tạo CSDL", str(e))
    finally:
        conn.close()

# ====== HÀM TIỆN ÍCH VÀ TÌM KIẾM ======
def center_window(win, w=1050, h=680): # Tăng kích thước cửa sổ
    """Căn giữa cửa sổ ứng dụng"""
    ws = win.winfo_screenwidth()
    hs = win.winfo_screenheight()
    x = (ws // 2) - (w // 2)
    y = (hs // 2) - (h // 2)
    win.geometry(f'{w}x{h}+{x}+{y}')

def clear_input():
    """Xóa trắng các ô nhập liệu"""
    entry_maso.delete(0, tk.END)
    entry_holot.delete(0, tk.END)
    entry_ten.delete(0, tk.END)
    gender_var.set("Nam")
    date_entry_ns.set_date("2000-01-01")
    entry_lop.delete(0, tk.END)
    entry_phong.delete(0, tk.END)
    # Các trường mới
    entry_thanhtien.delete(0, tk.END)
    entry_thanhtien.insert(0, "2000000") # Mặc định 2.000.000
    status_var.set("Chưa đóng")
    date_entry_nv.set_date(date.today().strftime("%Y-%m-%d"))
    entry_maso.config(state='normal')

def load_data():
    """Tải và hiển thị TOÀN BỘ dữ liệu sinh viên lên Treeview (có cột mới)"""
    for i in tree.get_children():
        tree.delete(i)
    conn = connect_db()
    try:
        cur = conn.cursor()
        # ĐÃ SỬA: Chuyển ngayvao ra sau lop
        cur.execute("""
            SELECT maso, holot, ten, gioitinh, ngaysinh, lop, ngayvao, phong_so, thanhtien, trangthaidongtien
            FROM sinhvien
            ORDER BY maso
        """)
        for row in cur.fetchall():
            row_list = list(row)
            # Định dạng lại ngày tháng
            if row_list[4]: # ngaysinh
                 row_list[4] = row_list[4].strftime("%Y-%m-%d")
            # row_list[6] là ngayvao theo thứ tự SELECT mới
            if row_list[6]: # ngayvao (vị trí 6)
                 row_list[6] = row_list[6].strftime("%Y-%m-%d")
            # Định dạng tiền tệ
            row_list[8] = f"{int(row_list[8]):,} VNĐ" if row_list[8] is not None else "" # thanhtien (vị trí 8)

            # Thêm row vào treeview
            tree.insert("", tk.END, values=row_list)
    except Exception as e:
        print(f"Lỗi tải dữ liệu: {e}")
    finally:
        conn.close()

def search_data():
    """Tìm kiếm và hiển thị dữ liệu sinh viên (cập nhật SQL)"""
    search_by = search_var.get()
    keyword = entry_search.get().strip()

    for i in tree.get_children():
        tree.delete(i)

    conn = connect_db()
    try:
        cur = conn.cursor()
        # ĐÃ SỬA: Chuyển ngayvao ra sau lop
        sql_query = """
            SELECT maso, holot, ten, gioitinh, ngaysinh, lop, ngayvao, phong_so, thanhtien, trangthaidongtien
            FROM sinhvien
        """
        params = ()

        if not keyword and search_by != "Tất cả":
             load_data()
             return

        if search_by == "Tất cả":
            if keyword:
                like_keyword = f"%{keyword}%"
                sql_query += """ WHERE maso LIKE %s OR holot LIKE %s OR ten LIKE %s
                                 OR lop LIKE %s OR phong_so LIKE %s OR trangthaidongtien LIKE %s"""
                params = (like_keyword, like_keyword, like_keyword, like_keyword, like_keyword, like_keyword)
            else:
                 load_data()
                 return

        elif search_by == "Mã SV" and keyword:
            sql_query += " WHERE maso LIKE %s"
            params = (f"%{keyword}%",)

        elif search_by == "Họ Tên" and keyword:
            like_keyword = f"%{keyword}%"
            sql_query += " WHERE holot LIKE %s OR ten LIKE %s OR CONCAT(holot, ' ', ten) LIKE %s"
            params = (like_keyword, like_keyword, like_keyword)

        elif search_by == "Lớp" and keyword:
            sql_query += " WHERE lop LIKE %s"
            params = (f"%{keyword}%",)

        elif search_by == "Phòng" and keyword:
            sql_query += " WHERE phong_so LIKE %s"
            params = (f"%{keyword}%",)

        elif search_by == "Trạng thái đóng tiền" and keyword:
            sql_query += " WHERE trangthaidongtien LIKE %s"
            params = (f"%{keyword}%",)

        sql_query += " ORDER BY maso"

        cur.execute(sql_query, params)
        rows = cur.fetchall()

        if not rows:
            messagebox.showinfo("Thông báo", "Không tìm thấy sinh viên nào phù hợp với từ khóa.")
            load_data()
            return

        for row in rows:
            row_list = list(row)
            if row_list[4]:
                row_list[4] = row_list[4].strftime("%Y-%m-%d")
            # row_list[6] là ngayvao theo thứ tự SELECT mới
            if row_list[6]:
                row_list[6] = row_list[6].strftime("%Y-%m-%d")
            row_list[8] = f"{int(row_list[8]):,} VNĐ" if row_list[8] is not None else "" # thanhtien (vị trí 8)
            tree.insert("", tk.END, values=row_list)

    except Exception as e:
        messagebox.showerror("Lỗi tìm kiếm", str(e))
    finally:
        conn.close()


def XuatExcel():
    """Xuất toàn bộ dữ liệu ra file Excel (cập nhật cột mới)"""
    conn = connect_db()
    try:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                                 title="Lưu file Excel")
        if file_path:
            cur = conn.cursor()
            # ĐÃ SỬA: Chuyển ngayvao ra sau lop
            cur.execute("""
                SELECT maso, lop, ngayvao, holot, ten, gioitinh, ngaysinh, phong_so, thanhtien, trangthaidongtien
                FROM sinhvien ORDER BY maso
            """)
            rows = cur.fetchall()

            wb = Workbook()
            ws = wb.active
            ws.title = "Danh Sách Sinh Viên Ở Ký Túc Xá"

            # ĐÃ SỬA: Cập nhật tiêu đề Excel theo thứ tự mới
            headers = ["Mã SV", "Lớp", "Ngày Vào", "Họ Lót", "Tên", "Giới Tính", "Ngày Sinh", "Phòng", "Thành Tiền", "Trạng Thái Đóng Tiền"]
            ws.append(headers)

            header_font = Font(bold=True, color="FFFFFF")
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = openpyxl.styles.PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            for row_data in rows:
                row_list = list(row_data)
                # Định dạng ngày tháng (Vị trí ngày sinh và ngày vào đã thay đổi)
                if row_list[6]:
                    row_list[6] = row_list[6].strftime("%Y-%m-%d") # Ngaysinh (vị trí 6)
                if row_list[2]:
                    row_list[2] = row_list[2].strftime("%Y-%m-%d") # Ngayvao (vị trí 2)
                ws.append(row_list)

                for cell in ws[ws.max_row]:
                    cell.border = thin_border

            # Tự động điều chỉnh độ rộng cột
            for column_cells in ws.columns:
                length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 5

            wb.save(file_path)
            messagebox.showinfo("Thành công", f"Đã xuất file Excel tại:\n{file_path}")

    except Exception as e:
        messagebox.showerror("Lỗi", str(e))
    finally:
        conn.close()

# ====== HÀM CRUD (Cập nhật) ======
def them_sv():
    """Thêm sinh viên mới (cập nhật SQL)"""
    maso = entry_maso.get().strip()
    holot = entry_holot.get().strip()
    ten = entry_ten.get().strip()
    gioitinh = gender_var.get()
    ngaysinh = date_entry_ns.get_date()
    lop = entry_lop.get().strip()
    phong = entry_phong.get().strip()
    # Các trường mới
    thanhtien_str = entry_thanhtien.get().strip().replace(" VNĐ", "").replace(",", "")
    try:
        thanhtien = int(thanhtien_str)
    except ValueError:
        messagebox.showwarning("Lỗi nhập liệu", "Thành tiền phải là một số nguyên.")
        return

    trangthai = status_var.get()
    ngayvao = date_entry_nv.get_date()


    if maso == "" or holot == "" or ten == "" or lop == "" or phong == "":
        messagebox.showwarning("Thiếu dữ liệu", "Vui lòng nhập đủ Mã SV, Họ tên, Lớp, Phòng và Thành tiền")
        return

    conn = connect_db()
    try:
        cur = conn.cursor()
        # Cập nhật SQL để bao gồm các cột mới
        sql = "INSERT INTO sinhvien (maso, holot, ten, gioitinh, ngaysinh, lop, phong_so, thanhtien, trangthaidongtien, ngayvao) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        val = (maso, holot, ten, gioitinh, ngaysinh, lop, phong, thanhtien, trangthai, ngayvao)
        cur.execute(sql, val)
        conn.commit()
        messagebox.showinfo("Thành công", f"Đã thêm sinh viên {holot} {ten}")
        load_data()
        clear_input()
    except mysql.connector.IntegrityError:
        messagebox.showerror("Lỗi", f"Mã sinh viên '{maso}' đã tồn tại!")
    except Exception as e:
        messagebox.showerror("Lỗi", str(e))
    finally:
        conn.close()

def xoa_sv():
    """Xóa sinh viên được chọn"""
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Chưa chọn", "Hãy chọn sinh viên để xóa")
        return

    # Lấy maso từ cột đầu tiên của dữ liệu
    maso = tree.item(selected)["values"][0]
    ten_sv = tree.item(selected)["values"][2]

    confirm = messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa sinh viên {ten_sv} có mã {maso}?")
    if confirm:
        conn = connect_db()
        try:
            cur = conn.cursor()
            cur.execute("DELETE FROM sinhvien WHERE maso=%s", (maso,))
            conn.commit()
            messagebox.showinfo("Thành công", "Đã xóa sinh viên")
            load_data()
            clear_input()
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
        finally:
            conn.close()

def sua_nv(event=None):
    """Tải thông tin sinh viên lên form để sửa (cập nhật)"""
    selected = tree.selection()
    if not selected:
        if event is None:
              messagebox.showwarning("Chưa chọn", "Hãy chọn sinh viên để sửa")
        return

    values = tree.item(selected)["values"]

    # THỨ TỰ values ĐÃ THAY ĐỔI do thay đổi SELECT trong load_data
    # Thứ tự mới: maso(0), holot(1), ten(2), gioitinh(3), ngaysinh(4), lop(5), ngayvao(6), phong_so(7), thanhtien(8), trangthaidongtien(9)

    # Chuyển đổi Thành Tiền từ chuỗi "X,XXX,XXX VNĐ" về số nguyên (vị trí 8)
    thanhtien_display = values[8].replace(" VNĐ", "").replace(",", "")

    clear_input()

    entry_maso.insert(0, values[0])
    entry_maso.config(state='readonly')

    entry_holot.insert(0, values[1])
    entry_ten.insert(0, values[2])
    gender_var.set(values[3])
    date_entry_ns.set_date(values[4]) # Ngày sinh (vị trí 4)
    entry_lop.insert(0, values[5])
    entry_phong.insert(0, values[7]) # Phòng (vị trí 7)
    # Các trường mới
    entry_thanhtien.delete(0, tk.END)
    entry_thanhtien.insert(0, thanhtien_display) # Thành tiền (vị trí 8)
    status_var.set(values[9]) # Trạng thái đóng tiền (vị trí 9)
    date_entry_nv.set_date(values[6]) # Ngày vào (vị trí 6)

def luu_nv():
    """Lưu thông tin đã sửa của sinh viên (cập nhật SQL)"""
    maso = entry_maso.get().strip()
    holot = entry_holot.get().strip()
    ten = entry_ten.get().strip()
    gioitinh = gender_var.get()
    ngaysinh = date_entry_ns.get_date()
    lop = entry_lop.get().strip()
    phong = entry_phong.get().strip()
    # Các trường mới
    thanhtien_str = entry_thanhtien.get().strip().replace(" VNĐ", "").replace(",", "")
    try:
        thanhtien = int(thanhtien_str)
    except ValueError:
        messagebox.showwarning("Lỗi nhập liệu", "Thành tiền phải là một số nguyên.")
        return
    trangthai = status_var.get()
    ngayvao = date_entry_nv.get_date()

    if maso == "" or entry_maso.cget('state') != 'readonly':
        messagebox.showwarning("Cảnh báo", "Vui lòng chọn Sửa một sinh viên trước khi Lưu")
        return

    conn = connect_db()
    try:
        cur = conn.cursor()
        # Cập nhật SQL để bao gồm các cột mới
        sql = """UPDATE sinhvien
                  SET holot=%s, ten=%s, gioitinh=%s, ngaysinh=%s, lop=%s, phong_so=%s,
                      thanhtien=%s, trangthaidongtien=%s, ngayvao=%s
                  WHERE maso=%s"""
        val = (holot, ten, gioitinh, ngaysinh, lop, phong, thanhtien, trangthai, ngayvao, maso)
        cur.execute(sql, val)
        conn.commit()
        messagebox.showinfo("Thành công", "Cập nhật thông tin sinh viên thành công")
        load_data()
        clear_input()
    except Exception as e:
        messagebox.showerror("Lỗi", str(e))
    finally:
        conn.close()

# ====== GIAO DIỆN TKINTER (Cập nhật) ======
root = tk.Tk()
root.title("Quản lý Ký túc xá (Cập nhật)")
center_window(root, 1050, 680)
root.resizable(True, True)

# Khởi tạo CSDL/Bảng
initialize_db()

# Tiêu đề
lbl_title = tk.Label(root, text="QUẢN LÝ KÝ TÚC XÁ", font=("Arial", 18, "bold"), fg="#2c3e50")
lbl_title.pack(pady=10)

# Frame nhập thông tin
frame_info = tk.Frame(root)
frame_info.pack(pady=5, padx=20, fill="x")

# --- Nhóm 1: Cơ bản (Hàng 0, 1) ---
# Hàng 0: Mã SV, Lớp, Phòng
tk.Label(frame_info, text="Mã SV:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
entry_maso = tk.Entry(frame_info, width=15)
entry_maso.grid(row=0, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame_info, text="Lớp:").grid(row=0, column=2, padx=10, pady=5, sticky="w")
entry_lop = tk.Entry(frame_info, width=15)
entry_lop.grid(row=0, column=3, padx=5, pady=5, sticky="w")

tk.Label(frame_info, text="Phòng:").grid(row=0, column=4, padx=10, pady=5, sticky="w")
entry_phong = tk.Entry(frame_info, width=15)
entry_phong.grid(row=0, column=5, padx=5, pady=5, sticky="w")

# Hàng 1: Họ lót, Tên, Ngày sinh
tk.Label(frame_info, text="Họ lót:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
entry_holot = tk.Entry(frame_info, width=25)
entry_holot.grid(row=1, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame_info, text="Tên:").grid(row=1, column=2, padx=10, pady=5, sticky="w")
entry_ten = tk.Entry(frame_info, width=15)
entry_ten.grid(row=1, column=3, padx=5, pady=5, sticky="w")

tk.Label(frame_info, text="Ngày sinh:").grid(row=1, column=4, padx=10, pady=5, sticky="w")
date_entry_ns = DateEntry(frame_info, width=12, background="darkblue",
                        foreground="white", date_pattern="yyyy-mm-dd")
date_entry_ns.grid(row=1, column=5, padx=5, pady=5, sticky="w")

# --- Nhóm 2: Thông tin KTX (Hàng 2) ---
# Hàng 2: Giới tính, Thành tiền, TT Đóng tiền, Ngày vào
tk.Label(frame_info, text="Giới tính:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
gender_var = tk.StringVar(value="Nam")
frame_gender = tk.Frame(frame_info)
frame_gender.grid(row=2, column=1, sticky="w", padx=5)
tk.Radiobutton(frame_gender, text="Nam", variable=gender_var, value="Nam").pack(side=tk.LEFT)
tk.Radiobutton(frame_gender, text="Nữ", variable=gender_var, value="Nữ").pack(side=tk.LEFT, padx=10)

tk.Label(frame_info, text="Thành tiền:").grid(row=2, column=2, padx=10, pady=5, sticky="w")
entry_thanhtien = tk.Entry(frame_info, width=15)
entry_thanhtien.grid(row=2, column=3, padx=5, pady=5, sticky="w")
entry_thanhtien.insert(0, "2000000") # Giá trị mặc định

tk.Label(frame_info, text="TT Đóng tiền:").grid(row=3, column=0, padx=5, pady=5, sticky="w")
status_var = tk.StringVar(value="Chưa đóng")
status_options = ["Chưa đóng", "Đã đóng", "Quá hạn"]
status_combo = ttk.Combobox(frame_info, textvariable=status_var, values=status_options, width=12, state="readonly")
status_combo.grid(row=3, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame_info, text="Ngày vào:").grid(row=3, column=2, padx=10, pady=5, sticky="w")
date_entry_nv = DateEntry(frame_info, width=12, background="darkblue",
                        foreground="white", date_pattern="yyyy-mm-dd")
date_entry_nv.grid(row=3, column=3, padx=5, pady=5, sticky="w")


# Frame nút chức năng CRUD
frame_btn = tk.Frame(root)
frame_btn.pack(pady=15)

btn_width = 10
tk.Button(frame_btn, text="Thêm", width=btn_width, command=them_sv, bg="#2ecc71", fg="white").grid(row=0, column=0, padx=5)
tk.Button(frame_btn, text="Lưu", width=btn_width, command=luu_nv, bg="#3498db", fg="white").grid(row=0, column=1, padx=5)
tk.Button(frame_btn, text="Sửa", width=btn_width, command=sua_nv, bg="#f1c40f").grid(row=0, column=2, padx=5)
tk.Button(frame_btn, text="Hủy", width=btn_width, command=clear_input, bg="#bdc3c7").grid(row=0, column=3, padx=5)
tk.Button(frame_btn, text="Xóa", width=btn_width, command=xoa_sv, bg="#e74c3c", fg="white").grid(row=0, column=4, padx=5)
tk.Button(frame_btn, text="Tải lại", width=btn_width, command=load_data, bg="#95a5a6", fg="white").grid(row=0, column=5, padx=5)
tk.Button(frame_btn, text="Xuất Excel", width=btn_width, command=XuatExcel, bg="#1abc9c", fg="white").grid(row=0, column=6, padx=5)
tk.Button(frame_btn, text="Thoát", width=btn_width, command=root.quit, bg="#7f8c8d").grid(row=0, column=7, padx=5)

# --- KHUNG TÌM KIẾM ---
frame_search = tk.Frame(root)
frame_search.pack(pady=5, padx=20, fill="x")

search_var = tk.StringVar()
search_var.set("Tất cả")

# Cập nhật tùy chọn tìm kiếm
search_options = ["Tất cả", "Mã SV", "Họ Tên", "Lớp", "Phòng", "Trạng thái đóng tiền"]
tk.Label(frame_search, text="Tìm theo:").grid(row=0, column=0, padx=5, sticky="w")
search_combo = ttk.Combobox(frame_search, textvariable=search_var, values=search_options, width=15, state="readonly")
search_combo.grid(row=0, column=1, padx=5, sticky="w")
search_combo.bind("<<ComboboxSelected>>", lambda e: entry_search.delete(0, tk.END))

tk.Label(frame_search, text="Từ khóa:").grid(row=0, column=2, padx=15, sticky="w")
entry_search = tk.Entry(frame_search, width=40)
entry_search.grid(row=0, column=3, padx=5, sticky="w")

# Thêm nút Tìm kiếm
tk.Button(frame_search, text="Tìm kiếm", command=search_data, bg="#3498db", fg="white").grid(row=0, column=4, padx=10)

# Bảng danh sách sinh viên
lbl_ds = tk.Label(root, text="Danh sách Sinh viên Ký túc xá", font=("Arial", 10, "bold"))
lbl_ds.pack(pady=5, anchor="w", padx=20)

# ĐÃ SỬA: Cập nhật cột cho Treeview (ngayvao ra sau lop)
columns = ("maso", "holot", "ten", "gioitinh", "ngaysinh", "lop", "ngayvao", "phong_so", "thanhtien", "trangthaidongtien")
tree = ttk.Treeview(root, columns=columns, show="headings", height=12)

# Thiết lập tiêu đề cột (Theo thứ tự mới)
tree.heading("maso", text="Mã SV")
tree.heading("holot", text="Họ lót")
tree.heading("ten", text="Tên")
tree.heading("gioitinh", text="GT")
tree.heading("ngaysinh", text="Ngày sinh")
tree.heading("lop", text="Lớp")
tree.heading("ngayvao", text="Ngày vào")   # Cột Ngày vào đã được chuyển lên
tree.heading("phong_so", text="Phòng")
tree.heading("thanhtien", text="Thành tiền")
tree.heading("trangthaidongtien", text="TT Đóng tiền")

# Thiết lập độ rộng cột (Theo thứ tự mới)
tree.column("maso", width=70, anchor="center")
tree.column("holot", width=120)
tree.column("ten", width=70)
tree.column("gioitinh", width=40, anchor="center")
tree.column("ngaysinh", width=90, anchor="center")
tree.column("lop", width=80)
tree.column("ngayvao", width=90, anchor="center") # Độ rộng cho Ngày vào
tree.column("phong_so", width=60, anchor="center")
tree.column("thanhtien", width=100, anchor="e")
tree.column("trangthaidongtien", width=100, anchor="center")

# Thêm Scrollbar
scrollbar_y = ttk.Scrollbar(root, orient="vertical", command=tree.yview)
scrollbar_x = ttk.Scrollbar(root, orient="horizontal", command=tree.xview)
tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

tree.pack(padx=20, pady=5, fill="both", expand=True)
scrollbar_y.pack(side="right", fill="y")
scrollbar_x.pack(side="bottom", fill="x")


# Gán sự kiện Double-click để sửa
tree.bind("<Double-1>", sua_nv)

# ====== LOAD DỮ LIỆU BAN ĐẦU VÀ CHẠY ỨNG DỤNG ======
load_data()
clear_input()
root.mainloop()